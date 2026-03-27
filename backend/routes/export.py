from flask import Blueprint, request, jsonify, send_file
from config import get_collections
from bson import ObjectId
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import os
from datetime import datetime

export_bp = Blueprint('export', __name__)

@export_bp.route('/api/export/excel', methods=['POST'])
def export_excel():
    data = request.get_json()
    results = data.get('results', [])
    
    if not results:
        return jsonify({'error': 'No results to export'}), 400
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resume Rankings"
    
    # Styles
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    tier1_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
    tier2_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
    tier3_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='e2e8f0'),
        right=Side(style='thin', color='e2e8f0'),
        top=Side(style='thin', color='e2e8f0'),
        bottom=Side(style='thin', color='e2e8f0')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f'AI Resume Classifier - Results ({datetime.now().strftime("%Y-%m-%d")})'
    ws['A1'].font = Font(size=14, bold=True, color="1e293b")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Rank', 'Candidate', 'Score', 'Skills %', 'Experience %', 'Education %', 'Tier']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data
    for i, r in enumerate(results):
        row = i + 4
        ws.cell(row=row, column=1, value=i + 1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=r['filename'])
        ws.cell(row=row, column=3, value=r['score']).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4, value=r['breakdown']['skills_score']).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=r['breakdown']['experience_score']).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=6, value=r['breakdown']['education_score']).alignment = Alignment(horizontal='center')
        
        tier_cell = ws.cell(row=row, column=7, value=r['tier'])
        tier_cell.alignment = Alignment(horizontal='center')
        
        if 'Tier 1' in r['tier']:
            tier_cell.fill = tier1_fill
        elif 'Tier 2' in r['tier']:
            tier_cell.fill = tier2_fill
        else:
            tier_cell.fill = tier3_fill
        
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 25
    
    filepath = os.path.join(os.path.dirname(__file__), '..', 'uploads', 'results.xlsx')
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, download_name='resume_rankings.xlsx')

@export_bp.route('/api/export/pdf', methods=['POST'])
def export_pdf():
    data = request.get_json()
    results = data.get('results', [])
    
    if not results:
        return jsonify({'error': 'No results to export'}), 400
    
    filepath = os.path.join(os.path.dirname(__file__), '..', 'uploads', 'results.pdf')
    
    doc = SimpleDocTemplate(filepath, pagesize=landscape(letter), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title = Paragraph(f'AI Resume Classifier - Results ({datetime.now().strftime("%Y-%m-%d")})', styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Table data
    table_data = [['Rank', 'Candidate', 'Score', 'Skills %', 'Exp %', 'Edu %', 'Tier']]
    
    for i, r in enumerate(results):
        table_data.append([
            str(i + 1),
            r['filename'],
            str(r['score']),
            str(r['breakdown']['skills_score']),
            str(r['breakdown']['experience_score']),
            str(r['breakdown']['education_score']),
            r['tier']
        ])
    
    table = Table(table_data, colWidths=[40, 180, 50, 60, 50, 50, 150])
    
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ])
    
    # Color tier cells
    for i, r in enumerate(results):
        row = i + 1
        if 'Tier 1' in r['tier']:
            style.add('BACKGROUND', (6, row), (6, row), colors.HexColor('#dcfce7'))
        elif 'Tier 2' in r['tier']:
            style.add('BACKGROUND', (6, row), (6, row), colors.HexColor('#fef3c7'))
        else:
            style.add('BACKGROUND', (6, row), (6, row), colors.HexColor('#fee2e2'))
    
    table.setStyle(style)
    elements.append(table)
    
    doc.build(elements)
    
    return send_file(filepath, as_attachment=True, download_name='resume_rankings.pdf')